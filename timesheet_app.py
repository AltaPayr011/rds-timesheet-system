import streamlit as st
import json
import os
from datetime import datetime, timedelta
from collections import defaultdict
import hashlib
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import base64

# File paths
EMPLOYEE_DATA_FILE = "employee_data.json"
USERS_FILE = "users.json"

# Load employee data
def load_employee_data():
    if os.path.exists(EMPLOYEE_DATA_FILE):
        with open(EMPLOYEE_DATA_FILE, 'r') as f:
            return json.load(f)
    return {}

# Save employee data
def save_employee_data(employee_data):
    with open(EMPLOYEE_DATA_FILE, 'w') as f:
        json.dump(employee_data, f, indent=2)

# Load user data
def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    return {
        "admin": {
            "password": hashlib.sha256("admin123".encode()).hexdigest(),
            "is_admin": True
        }
    }

# Save user data
def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

# Hash password
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# Verify password
def verify_password(username, password, users):
    if username in users:
        return users[username]["password"] == hash_password(password)
    return False

# Convert image to base64
def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

# Find employee key by full name (handles multi-word surnames, case variations, spelling)
def find_employee_key(full_name, employee_data):
    # Try exact match first
    test_key = full_name.replace(" ", "_")
    if test_key in employee_data:
        return test_key
    
    # Normalize the input name for comparison
    full_name_lower = full_name.lower()
    
    # Try case-insensitive matching
    for key, emp in employee_data.items():
        emp_full_name = f"{emp['first_name']} {emp['last_name']}"
        if emp_full_name.lower() == full_name_lower:
            return key
    
    # Try fuzzy matching for common variations (Jansen/Janse, van/Van)
    for key, emp in employee_data.items():
        emp_full_name = f"{emp['first_name']} {emp['last_name']}"
        emp_full_name_normalized = emp_full_name.lower().replace("jansen", "janse").replace("janse", "janse")
        full_name_normalized = full_name_lower.replace("jansen", "janse").replace("janse", "janse")
        
        if emp_full_name_normalized == full_name_normalized:
            return key
    
    return None

# Login page
def login_page():
    if os.path.exists("RDS_Logo.jpg"):
        logo_base64 = get_base64_image("RDS_Logo.jpg")
        st.markdown(
            f'<div style="text-align: center;"><img src="data:image/jpeg;base64,{logo_base64}" width="200"></div>',
            unsafe_allow_html=True
        )
    
    st.markdown("<h1 style='text-align: center;'>RDS Timesheet System</h1>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.subheader("Login")
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        
        if st.button("Login", use_container_width=True):
            users = load_users()
            if verify_password(username, password, users):
                st.session_state['logged_in'] = True
                st.session_state['username'] = username
                st.session_state['is_admin'] = users[username].get('is_admin', False)
                st.rerun()
            else:
                st.error("Invalid username or password")

# Change password page
def change_password_page():
    st.title("🔒 Change Password")
    
    st.write(f"**Logged in as:** {st.session_state['username']}")
    
    users = load_users()
    
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.subheader("Change Your Password")
        
        current_password = st.text_input("Current Password", type="password", key="current_pwd")
        new_password = st.text_input("New Password", type="password", key="new_pwd")
        confirm_password = st.text_input("Confirm New Password", type="password", key="confirm_pwd")
        
        st.markdown("**Password Requirements:**")
        st.info("• Minimum 6 characters recommended\n• Use a mix of letters and numbers\n• Don't share your password with others")
        
        if st.button("Change Password", type="primary", use_container_width=True):
            # Validate current password
            if not verify_password(st.session_state['username'], current_password, users):
                st.error("❌ Current password is incorrect")
            elif not new_password:
                st.error("❌ Please enter a new password")
            elif len(new_password) < 6:
                st.warning("⚠️ Password should be at least 6 characters long")
            elif new_password != confirm_password:
                st.error("❌ New passwords do not match")
            elif new_password == current_password:
                st.warning("⚠️ New password must be different from current password")
            else:
                # Update password
                users[st.session_state['username']]["password"] = hash_password(new_password)
                save_users(users)
                st.success("✅ Password changed successfully!")
                st.balloons()
                
                # Show download option for Streamlit Cloud persistence
                st.markdown("---")
                st.info("💡 **For Streamlit Cloud users:** Download the updated users.json file and update it in your GitHub repository to make this change permanent.")
                
                import json
                users_json = json.dumps(users, indent=2)
                st.download_button(
                    label="📥 Download users.json",
                    data=users_json,
                    file_name="users.json",
                    mime="application/json"
                )
                
                st.markdown("---")
                st.info("💡 Please use your new password the next time you log in.")

# User management page
def user_management_page():
    st.title("👥 User Management")
    
    users = load_users()
    
    st.subheader("Create New User")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        new_username = st.text_input("New Username")
    with col2:
        new_password = st.text_input("New Password", type="password")
    with col3:
        is_admin = st.checkbox("Admin User")
    
    if st.button("Create User"):
        if new_username and new_password:
            if new_username in users:
                st.error(f"User '{new_username}' already exists")
            else:
                users[new_username] = {
                    "password": hash_password(new_password),
                    "is_admin": is_admin
                }
                save_users(users)
                st.success(f"User '{new_username}' created successfully")
                st.rerun()
        else:
            st.error("Please enter both username and password")
    
    st.markdown("---")
    st.subheader("Existing Users")
    
    user_list = []
    for username, user_data in users.items():
        user_list.append({
            "Username": username,
            "Role": "Admin" if user_data.get('is_admin', False) else "User"
        })
    
    if user_list:
        st.table(user_list)
    
    st.markdown("---")
    st.subheader("Reset Password")
    
    col1, col2 = st.columns(2)
    with col1:
        reset_username = st.selectbox("Select User", list(users.keys()))
    with col2:
        reset_password = st.text_input("New Password", type="password", key="reset_pass")
    
    if st.button("Reset Password"):
        if reset_password:
            users[reset_username]["password"] = hash_password(reset_password)
            save_users(users)
            st.success(f"Password reset for '{reset_username}'")
        else:
            st.error("Please enter a new password")

# Employee management page
def employee_management_page():
    st.title("👤 Employee Management")
    st.write("*Add, edit, or update employee information and working hours*")
    
    employee_data = load_employee_data()
    
    tab1, tab2, tab3 = st.tabs(["➕ Add Employee", "✏️ Edit Employee", "🗑️ Delete Employee"])
    
    with tab1:
        st.subheader("Add New Employee")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            first_name = st.text_input("First Name*")
        with col2:
            last_name = st.text_input("Last Name*")
        with col3:
            employee_number = st.text_input("Employee Number*")
        
        area = st.text_input("Area/Department")
        
        st.markdown("**Working Hours per Day:**")
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            mon_hours = st.number_input("Monday", min_value=0.0, max_value=24.0, value=8.75, step=0.25)
        with col2:
            tue_hours = st.number_input("Tuesday", min_value=0.0, max_value=24.0, value=8.75, step=0.25)
        with col3:
            wed_hours = st.number_input("Wednesday", min_value=0.0, max_value=24.0, value=8.75, step=0.25)
        with col4:
            thu_hours = st.number_input("Thursday", min_value=0.0, max_value=24.0, value=8.75, step=0.25)
        with col5:
            fri_hours = st.number_input("Friday", min_value=0.0, max_value=24.0, value=5.0, step=0.25)
        
        notes = st.text_area("Notes (Optional)")
        
        if st.button("Add Employee", type="primary"):
            if first_name and last_name and employee_number:
                emp_key = f"{first_name}_{last_name}"
                
                if emp_key in employee_data:
                    st.error(f"Employee '{first_name} {last_name}' already exists")
                else:
                    employee_data[emp_key] = {
                        "first_name": first_name,
                        "last_name": last_name,
                        "employee_number": employee_number,
                        "area": area,
                        "required_hours": {
                            "Monday": mon_hours,
                            "Tuesday": tue_hours,
                            "Wednesday": wed_hours,
                            "Thursday": thu_hours,
                            "Friday": fri_hours,
                            "Saturday": 0.0,
                            "Sunday": 0.0
                        },
                        "notes": notes
                    }
                    save_employee_data(employee_data)
                    st.success(f"✅ Employee '{first_name} {last_name}' added successfully")
                    
                    # Show download option for Streamlit Cloud persistence
                    st.info("💡 **For Streamlit Cloud users:** Download the updated employee_data.json file and update it in your GitHub repository to make this change permanent.")
                    
                    import json
                    emp_json = json.dumps(employee_data, indent=2)
                    st.download_button(
                        label="📥 Download employee_data.json",
                        data=emp_json,
                        file_name="employee_data.json",
                        mime="application/json",
                        key="download_after_add"
                    )
            else:
                st.error("Please fill in all required fields (marked with *)")
    
    with tab2:
        st.subheader("Edit Existing Employee")
        
        if not employee_data:
            st.info("No employees in database")
        else:
            emp_names = [f"{emp['first_name']} {emp['last_name']}" for emp in employee_data.values()]
            emp_keys = list(employee_data.keys())
            
            selected_name = st.selectbox("Select Employee to Edit", emp_names)
            selected_key = emp_keys[emp_names.index(selected_name)]
            emp = employee_data[selected_key]
            
            st.markdown("---")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                edit_first_name = st.text_input("First Name*", value=emp['first_name'], key="edit_first")
            with col2:
                edit_last_name = st.text_input("Last Name*", value=emp['last_name'], key="edit_last")
            with col3:
                edit_employee_number = st.text_input("Employee Number*", value=emp['employee_number'], key="edit_num")
            
            edit_area = st.text_input("Area/Department", value=emp.get('area', ''), key="edit_area")
            
            st.markdown("**Working Hours per Day:**")
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                edit_mon = st.number_input("Monday", min_value=0.0, max_value=24.0, 
                                          value=emp['required_hours']['Monday'], step=0.25, key="edit_mon")
            with col2:
                edit_tue = st.number_input("Tuesday", min_value=0.0, max_value=24.0, 
                                          value=emp['required_hours']['Tuesday'], step=0.25, key="edit_tue")
            with col3:
                edit_wed = st.number_input("Wednesday", min_value=0.0, max_value=24.0, 
                                          value=emp['required_hours']['Wednesday'], step=0.25, key="edit_wed")
            with col4:
                edit_thu = st.number_input("Thursday", min_value=0.0, max_value=24.0, 
                                          value=emp['required_hours']['Thursday'], step=0.25, key="edit_thu")
            with col5:
                edit_fri = st.number_input("Friday", min_value=0.0, max_value=24.0, 
                                          value=emp['required_hours']['Friday'], step=0.25, key="edit_fri")
            
            edit_notes = st.text_area("Notes (Optional)", value=emp.get('notes', ''), key="edit_notes")
            
            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("💾 Save Changes", type="primary"):
                    new_key = f"{edit_first_name}_{edit_last_name}"
                    if new_key != selected_key:
                        del employee_data[selected_key]
                    
                    employee_data[new_key] = {
                        "first_name": edit_first_name,
                        "last_name": edit_last_name,
                        "employee_number": edit_employee_number,
                        "area": edit_area,
                        "required_hours": {
                            "Monday": edit_mon,
                            "Tuesday": edit_tue,
                            "Wednesday": edit_wed,
                            "Thursday": edit_thu,
                            "Friday": edit_fri,
                            "Saturday": 0.0,
                            "Sunday": 0.0
                        },
                        "notes": edit_notes
                    }
                    save_employee_data(employee_data)
                    st.success(f"✅ Employee '{edit_first_name} {edit_last_name}' updated successfully")
                    
                    # Show download option for Streamlit Cloud persistence
                    st.info("💡 **For Streamlit Cloud users:** Download the updated employee_data.json file and update it in your GitHub repository to make this change permanent.")
                    
                    import json
                    emp_json = json.dumps(employee_data, indent=2)
                    st.download_button(
                        label="📥 Download employee_data.json",
                        data=emp_json,
                        file_name="employee_data.json",
                        mime="application/json",
                        key="download_after_edit"
                    )
    
    with tab3:
        st.subheader("Delete Employee")
        st.warning("⚠️ This action cannot be undone!")
        
        if not employee_data:
            st.info("No employees in database")
        else:
            emp_names = [f"{emp['first_name']} {emp['last_name']}" for emp in employee_data.values()]
            emp_keys = list(employee_data.keys())
            
            delete_name = st.selectbox("Select Employee to Delete", emp_names, key="delete_select")
            delete_key = emp_keys[emp_names.index(delete_name)]
            
            emp = employee_data[delete_key]
            
            st.markdown("---")
            st.markdown("**Employee Details:**")
            st.write(f"**Name:** {emp['first_name']} {emp['last_name']}")
            st.write(f"**Employee Number:** {emp['employee_number']}")
            st.write(f"**Area:** {emp.get('area', 'N/A')}")
            st.write(f"**Weekly Hours:** {sum([emp['required_hours'].get(day, 0) for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']])}")
            
            confirm_text = st.text_input("Type 'DELETE' to confirm deletion")
            
            if st.button("🗑️ Delete Employee", type="primary"):
                if confirm_text == "DELETE":
                    del employee_data[delete_key]
                    save_employee_data(employee_data)
                    st.success(f"✅ Employee '{delete_name}' deleted successfully")
                    
                    # Show download option for Streamlit Cloud persistence
                    st.info("💡 **For Streamlit Cloud users:** Download the updated employee_data.json file and update it in your GitHub repository to make this change permanent.")
                    
                    import json
                    emp_json = json.dumps(employee_data, indent=2)
                    st.download_button(
                        label="📥 Download employee_data.json",
                        data=emp_json,
                        file_name="employee_data.json",
                        mime="application/json",
                        key="download_after_delete"
                    )
                else:
                    st.error("Please type 'DELETE' to confirm")
    
    st.markdown("---")
    st.subheader(f"Current Employees ({len(employee_data)})")
    
    if employee_data:
        emp_list = []
        for key, emp in sorted(employee_data.items(), key=lambda x: x[1]['last_name']):
            weekly_hours = sum([emp['required_hours'].get(day, 0) for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']])
            emp_list.append({
                "Name": f"{emp['first_name']} {emp['last_name']}",
                "Employee #": emp['employee_number'],
                "Area": emp.get('area', 'N/A'),
                "Weekly Hours": f"{weekly_hours:.2f}h"
            })
        
        st.dataframe(emp_list, use_container_width=True)

# Parse HTML timesheet file
def parse_html_timesheet(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    
    timesheet_data = defaultdict(list)
    rows = soup.find_all('tr', {'index': True})
    
    for row in rows:
        divs = row.find_all('div', style=lambda x: x and 'float:left' in x)
        
        if len(divs) >= 8:
            employee = divs[1].text.strip()
            date_str = divs[2].text.strip()
            time_in = divs[3].text.strip()
            time_out = divs[4].text.strip()
            hours_str = divs[5].text.strip()
            job_type = divs[6].text.strip()
            shift = divs[7].text.strip()
            
            date_match = re.match(r'(\d+)/(\d+)', date_str)
            if date_match:
                day = int(date_match.group(1))
                month = int(date_match.group(2))
                year = 2025 if month == 12 else 2026
                
                date_obj = datetime(year, month, day)
                day_name = date_obj.strftime("%A")
                
                try:
                    hours = float(hours_str) if hours_str else 0.0
                except:
                    hours = 0.0
                
                timesheet_data[employee].append({
                    'date': date_obj,
                    'date_str': date_obj.strftime("%Y/%m/%d"),
                    'day': day_name,
                    'time_in': time_in if time_in else '',
                    'time_out': time_out if time_out else '',
                    'hours': hours,
                    'job_type': job_type,
                    'shift': shift
                })
    
    return timesheet_data

# Calculate required hours based on actual dates in the period
def calculate_required_hours(employee_key, dates, employee_data):
    if not employee_key or employee_key not in employee_data:
        # Default schedule
        required = 0
        for date_obj, day_name in dates:
            if day_name in ["Monday", "Tuesday", "Wednesday", "Thursday"]:
                required += 8.75
            elif day_name == "Friday":
                required += 5.0
        return required
    
    emp = employee_data[employee_key]
    required = 0
    
    for date_obj, day_name in dates:
        required += emp['required_hours'].get(day_name, 0.0)
    
    return required

# Apply OT/Call Out rounding rule
def apply_ot_rounding(entries):
    """
    For Overtime 2.0 and Call Out 2.0:
    If hours < 4 on a specific day, round UP to 4 for that day
    """
    # Group by date and job type
    by_date_and_type = defaultdict(float)
    
    for entry in entries:
        if 'Overtime 2.0' in entry['job_type'] or 'Call Out 2.0' in entry['job_type']:
            key = (entry['date_str'], entry['job_type'])
            by_date_and_type[key] += entry['hours']
    
    # Calculate total with rounding
    totals_by_type = defaultdict(float)
    
    for (date, job_type), hours in by_date_and_type.items():
        if hours > 0 and hours < 4:
            totals_by_type[job_type] += 4.0  # Round up to 4
        else:
            totals_by_type[job_type] += hours
    
    return totals_by_type

# Generate Excel report
def generate_excel_report(timesheet_data, employee_data, skip_unknown):
    wb = Workbook()
    detail_sheet = wb.active
    detail_sheet.title = "Timesheet Detail"
    
    # Headers
    headers = ["Employee Name", "Employee #", "Date", "Day of Week", "Time In", "Time Out", "Hours", "Job/Absence Type", "Shift", "Required Hours", "Short Time"]
    for col, header in enumerate(headers, 1):
        cell = detail_sheet.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Column widths
    detail_sheet.column_dimensions['A'].width = 30
    detail_sheet.column_dimensions['B'].width = 12
    detail_sheet.column_dimensions['C'].width = 12
    detail_sheet.column_dimensions['D'].width = 12
    detail_sheet.column_dimensions['E'].width = 10
    detail_sheet.column_dimensions['F'].width = 10
    detail_sheet.column_dimensions['G'].width = 10
    detail_sheet.column_dimensions['H'].width = 30
    detail_sheet.column_dimensions['I'].width = 25
    detail_sheet.column_dimensions['J'].width = 14
    detail_sheet.column_dimensions['K'].width = 12
    
    # Style fills
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    summary_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # Get date range
    all_dates = []
    for entries in timesheet_data.values():
        all_dates.extend([e['date'] for e in entries])
    
    start_date = min(all_dates)
    end_date = max(all_dates)
    
    # Generate full date range
    current_date = start_date
    dates = []
    while current_date <= end_date:
        dates.append((current_date, current_date.strftime("%A")))
        current_date += timedelta(days=1)
    
    row = 2
    
    # Process each employee
    for emp_name in sorted(timesheet_data.keys()):
        # Find employee key
        emp_key = find_employee_key(emp_name, employee_data)
        
        if skip_unknown and not emp_key:
            continue
        
        # Get employee number and schedule info
        emp_number = ""
        emp_first_name = ""
        if emp_key and emp_key in employee_data:
            emp_number = employee_data[emp_key].get('employee_number', '')
            emp_first_name = employee_data[emp_key].get('first_name', '')
        
        entries = timesheet_data[emp_name]
        
        # Create date lookup and calculate hours per date
        entry_by_date = {}
        hours_by_date = defaultdict(float)
        
        for entry in entries:
            date_str = entry['date_str']
            if date_str not in entry_by_date:
                entry_by_date[date_str] = []
            entry_by_date[date_str].append(entry)
            hours_by_date[date_str] += entry['hours']
        
        # Check if this is Eunice or Veronica for special filtering
        is_eunice = "Eunice" in emp_name
        is_veronica = "Veronica" in emp_name
        
        # Process all dates
        for date_obj, day_name in dates:
            date_str = date_obj.strftime("%Y/%m/%d")
            
            # Calculate required hours for this specific day
            required_for_day = 0.0
            if emp_key and emp_key in employee_data:
                required_for_day = employee_data[emp_key]['required_hours'].get(day_name, 0.0)
            else:
                # Default schedule
                if day_name in ["Monday", "Tuesday", "Wednesday", "Thursday"]:
                    required_for_day = 8.75
                elif day_name == "Friday":
                    required_for_day = 5.0
            
            # Calculate actual hours worked this day
            actual_hours_day = hours_by_date.get(date_str, 0.0)
            
            # Calculate short time for this day
            short_time_day = max(0, required_for_day - actual_hours_day)
            
            # Skip certain days for specific employees if no transactions
            if date_str not in entry_by_date:
                if is_eunice and day_name in ["Tuesday", "Wednesday", "Friday"]:
                    continue  # Skip these days for Eunice if no transactions
                if is_veronica and day_name == "Friday":
                    continue  # Skip Fridays for Veronica if no transactions
            
            if date_str in entry_by_date:
                # Show required hours and short time only on first entry for this date
                first_entry = True
                for entry in entry_by_date[date_str]:
                    detail_sheet.cell(row, 1, emp_name)
                    detail_sheet.cell(row, 2, emp_number)
                    detail_sheet.cell(row, 3, date_str)
                    detail_sheet.cell(row, 4, entry['day'])
                    detail_sheet.cell(row, 5, entry['time_in'])
                    detail_sheet.cell(row, 6, entry['time_out'])
                    detail_sheet.cell(row, 7, entry['hours'])
                    detail_sheet.cell(row, 8, entry['job_type'])
                    detail_sheet.cell(row, 9, entry['shift'])
                    
                    # Show required hours and short time only on first row for this date
                    if first_entry:
                        if required_for_day > 0:
                            detail_sheet.cell(row, 10, round(required_for_day, 2))
                        if short_time_day > 0:
                            detail_sheet.cell(row, 11, round(short_time_day, 2))
                    
                    first_entry = False
                    row += 1
            else:
                # Grey row for missing date
                detail_sheet.cell(row, 1, emp_name).fill = grey_fill
                detail_sheet.cell(row, 2, emp_number).fill = grey_fill
                detail_sheet.cell(row, 3, date_str).fill = grey_fill
                detail_sheet.cell(row, 4, day_name).fill = grey_fill
                for col in range(5, 10):
                    detail_sheet.cell(row, col).fill = grey_fill
                
                # Show required hours and short time for missing days
                if required_for_day > 0:
                    detail_sheet.cell(row, 10, round(required_for_day, 2)).fill = grey_fill
                if short_time_day > 0:
                    detail_sheet.cell(row, 11, round(short_time_day, 2)).fill = grey_fill
                
                row += 1
        
        # Calculate summary
        job_totals = defaultdict(float)
        for entry in entries:
            job_totals[entry['job_type']] += entry['hours']
        
        # Add summary rows
        detail_sheet.cell(row, 1, f"{emp_name} - SUMMARY").font = Font(bold=True)
        detail_sheet.cell(row, 1).fill = summary_fill
        for col in range(2, 12):
            detail_sheet.cell(row, col).fill = summary_fill
        row += 1
        
        for job_type in sorted(job_totals.keys()):
            detail_sheet.cell(row, 3, job_type).fill = summary_fill
            detail_sheet.cell(row, 7, round(job_totals[job_type], 2)).fill = summary_fill
            for col in [1, 2, 4, 5, 6, 8, 9, 10, 11]:
                detail_sheet.cell(row, col).fill = summary_fill
            row += 1
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Employee Summary")
    
    # Collect all job types
    all_job_types = set()
    for entries in timesheet_data.values():
        for entry in entries:
            all_job_types.add(entry['job_type'])
    
    # Define standard columns
    standard_columns = ["Normal Working Hours", "Paid Leave - Any type", "Unpaid Leave - Any Type"]
    
    # Identify OT columns (excluding Overtime 2.0 and Call Out 2.0 which have special rounding)
    ot_special = ["Overtime 2.0", "Call Out 2.0"]
    ot_columns = [jt for jt in all_job_types 
                  if ('OT' in jt or 'Overtime' in jt or 'Call Out' in jt) 
                  and jt not in ot_special]
    
    # Other columns
    other_columns = [jt for jt in all_job_types 
                    if jt not in standard_columns 
                    and jt not in ot_columns
                    and jt not in ot_special
                    and 'Normal Working' not in jt
                    and 'Paid Leave' not in jt
                    and 'Unpaid Leave' not in jt]
    
    # Build headers
    summary_headers = ["Employee Name", "Employee #", "Normal Working Hours", "Paid Leave", "Unpaid Leave", 
                       "Total Regular Hours", "Required Hours", "Short Time"]
    
    # Add special OT columns first
    for ot_col in sorted(ot_special):
        if ot_col in all_job_types:
            summary_headers.append(ot_col)
    
    # Add regular OT columns
    for ot_col in sorted(ot_columns):
        summary_headers.append(ot_col)
    
    # Add other columns
    for other_col in sorted(other_columns):
        summary_headers.append(other_col)
    
    # Write headers
    for col, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Set column widths
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 12
    for i in range(3, len(summary_headers) + 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i-1)//26) + chr(65 + (i-1)%26)
        try:
            summary_sheet.column_dimensions[col_letter].width = 18
        except:
            pass
    
    summary_row = 2
    
    for emp_name in sorted(timesheet_data.keys()):
        emp_key = find_employee_key(emp_name, employee_data)
        
        if skip_unknown and not emp_key:
            continue
        
        # Get employee number
        emp_number = ""
        if emp_key and emp_key in employee_data:
            emp_number = employee_data[emp_key].get('employee_number', '')
        
        entries = timesheet_data[emp_name]
        
        # Calculate totals for each job type
        job_totals = defaultdict(float)
        for entry in entries:
            job_totals[entry['job_type']] += entry['hours']
        
        # Apply rounding for Overtime 2.0 and Call Out 2.0
        rounded_totals = apply_ot_rounding(entries)
        
        # Update job_totals with rounded values
        for job_type, rounded_value in rounded_totals.items():
            job_totals[job_type] = rounded_value
        
        # Calculate standard columns
        normal_hours = sum(hours for jt, hours in job_totals.items() if 'Normal Working' in jt)
        paid_leave = sum(hours for jt, hours in job_totals.items() if 'Paid Leave' in jt)
        unpaid_leave = sum(hours for jt, hours in job_totals.items() if 'Unpaid Leave' in jt)
        
        # Total Regular = Normal + Paid Leave + Unpaid Leave
        total_regular = normal_hours + paid_leave + unpaid_leave
        
        # Calculate required hours based on employee schedule
        required_hours = calculate_required_hours(emp_key, dates, employee_data)
        short_time = max(0, required_hours - total_regular)
        
        # Write data
        col_idx = 1
        summary_sheet.cell(summary_row, col_idx, emp_name)
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, emp_number)
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, round(normal_hours, 2) if normal_hours > 0 else "")
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, round(paid_leave, 2) if paid_leave > 0 else "")
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, round(unpaid_leave, 2) if unpaid_leave > 0 else "")
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, round(total_regular, 2))
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, round(required_hours, 2))
        col_idx += 1
        summary_sheet.cell(summary_row, col_idx, round(short_time, 2) if short_time > 0 else "")
        col_idx += 1
        
        # Write special OT columns (with rounding)
        for ot_col in sorted(ot_special):
            if ot_col in all_job_types:
                hours = job_totals.get(ot_col, 0)
                summary_sheet.cell(summary_row, col_idx, round(hours, 2) if hours > 0 else "")
                col_idx += 1
        
        # Write regular OT columns
        for ot_col in sorted(ot_columns):
            hours = job_totals.get(ot_col, 0)
            summary_sheet.cell(summary_row, col_idx, round(hours, 2) if hours > 0 else "")
            col_idx += 1
        
        # Write other columns
        for other_col in sorted(other_columns):
            hours = job_totals.get(other_col, 0)
            summary_sheet.cell(summary_row, col_idx, round(hours, 2) if hours > 0 else "")
            col_idx += 1
        
        summary_row += 1
    
    return wb

# Main application
def main_app():
    st.set_page_config(page_title="RDS Timesheet System", page_icon="📊", layout="wide")
    
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False
    
    if not st.session_state['logged_in']:
        login_page()
        return
    
    # Sidebar
    with st.sidebar:
        if os.path.exists("RDS_Logo.jpg"):
            st.image("RDS_Logo.jpg", width=150)
        
        st.title("Navigation")
        st.write(f"**User:** {st.session_state['username']}")
        st.write(f"**Role:** {'Admin' if st.session_state.get('is_admin', False) else 'User'}")
        
        # Employee Management available to all users
        menu_options = ["📊 Generate Report", "👤 Employee Management", "🔒 Change Password"]
        
        # User Management only for admins
        if st.session_state.get('is_admin', False):
            menu_options.append("👥 User Management")
        
        page = st.radio("Menu", menu_options)
        
        if st.button("Logout"):
            st.session_state['logged_in'] = False
            st.session_state['username'] = None
            st.session_state['is_admin'] = False
            st.rerun()
    
    # Main content
    if page == "👥 User Management":
        user_management_page()
    elif page == "👤 Employee Management":
        employee_management_page()
    elif page == "🔒 Change Password":
        change_password_page()
    else:
        generate_report_page()

# Generate report page
def generate_report_page():
    st.title("📊 RDS Timesheet Report Generator")
    
    employee_data = load_employee_data()
    
    st.write(f"**Employees in Database:** {len(employee_data)}")
    
    st.subheader("1. Upload HTML Timesheet File")
    st.info("💡 **How to export from OpenTimeClock:**\n"
            "1. Go to **List View**\n"
            "2. Select your **date range** (start and end dates)\n"
            "3. Select **All Users**\n"
            "4. File → Save Page As → **Webpage, Complete**\n"
            "5. Upload the HTML file here")
    
    uploaded_file = st.file_uploader("Upload HTML file from OpenTimeClock", type=['html'])
    
    if uploaded_file:
        html_content = uploaded_file.read().decode('utf-8')
        timesheet_data = parse_html_timesheet(html_content)
        
        if not timesheet_data:
            st.error("No timesheet data found in HTML file")
            return
        
        st.success(f"✅ Parsed {len(timesheet_data)} employees with {sum(len(v) for v in timesheet_data.values())} total entries")
        
        all_dates = []
        for entries in timesheet_data.values():
            all_dates.extend([e['date'] for e in entries])
        
        start_date = min(all_dates)
        end_date = max(all_dates)
        st.write(f"**Date Range:** {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        
        unknown_employees = []
        for emp_name in timesheet_data.keys():
            if not find_employee_key(emp_name, employee_data):
                unknown_employees.append(emp_name)
        
        st.subheader("2. Report Options")
        
        skip_unknown = False
        if unknown_employees:
            st.warning(f"⚠️ Found {len(unknown_employees)} employees not in database:")
            for emp in unknown_employees[:5]:
                st.write(f"  • {emp}")
            if len(unknown_employees) > 5:
                st.write(f"  ... and {len(unknown_employees) - 5} more")
            
            skip_unknown = st.checkbox("Skip employees not in database", value=False)
        
        st.subheader("3. Generate Report")
        
        if st.button("Generate Excel Report", type="primary"):
            with st.spinner("Generating report..."):
                wb = generate_excel_report(timesheet_data, employee_data, skip_unknown)
                
                from io import BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                filename = f"Timesheet_Report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="📥 Download Excel Report",
                    data=buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("✅ Report generated successfully!")

if __name__ == "__main__":
    main_app()
